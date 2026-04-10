"""Excel workbook generator for delivery items."""

from __future__ import annotations

import io
import math
from dataclasses import dataclass
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from artifacts import DeliveryItem
except ModuleNotFoundError:
    from backend.artifacts import DeliveryItem


BOOST_KEYWORDS = [
    "core", "security", "auth", "foundation", "payment", "api",
    "database", "migration", "infrastructure", "audit", "workflow",
]
PRIORITY_TIERS = (
    ("P0 - Immediate", 70),
    ("P1 - Next", 56),
    ("P2 - Planned", 0),
)
PRIORITY_RANK = {tier: index for index, (tier, _) in enumerate(PRIORITY_TIERS)}
HALF_DAY_SLOTS = 2

PALETTE = {
    "navy": "17324D",
    "teal": "1F7A8C",
    "teal_soft": "DCEFF2",
    "gold": "E3A008",
    "gold_soft": "FEF3C7",
    "slate": "475467",
    "slate_soft": "F2F4F7",
    "mint": "0F766E",
    "mint_soft": "CCFBF1",
    "rose": "B42318",
    "rose_soft": "FEE4E2",
    "amber_soft": "FFF4CC",
    "white": "FFFFFF",
    "border": "D0D5DD",
    "text_muted": "475467",
}
PRIORITY_FILL = {
    "P0 - Immediate": PALETTE["gold_soft"],
    "P1 - Next": PALETTE["mint_soft"],
    "P2 - Planned": PALETTE["slate_soft"],
}
TRACEABILITY_FILL = {
    "explicit": PALETTE["mint_soft"],
    "inferred": PALETTE["amber_soft"],
    "unmapped": PALETTE["rose_soft"],
}
TRACEABILITY_LABEL = {
    "explicit": "Explicit",
    "inferred": "Inferred",
    "unmapped": "Unmapped",
}
THIN_BORDER = Border(
    left=Side(style="thin", color=PALETTE["border"]),
    right=Side(style="thin", color=PALETTE["border"]),
    top=Side(style="thin", color=PALETTE["border"]),
    bottom=Side(style="thin", color=PALETTE["border"]),
)


@dataclass
class PlanningRow:
    task_id: str
    priority_tier: str
    priority_score: int
    scheduling_bucket: str
    group: str
    title: str
    senior_rd_days: float
    tracker_estimate: int
    requirement_refs: list[str]
    requirement_source: str
    labels_text: str
    source_story: str
    slot_start: int
    slot_end: int

    @property
    def requirement_refs_text(self) -> str:
        return ", ".join(self.requirement_refs) if self.requirement_refs else "unmapped"

    @property
    def requirement_source_text(self) -> str:
        return TRACEABILITY_LABEL.get(self.requirement_source, self.requirement_source.title())

    @property
    def start_label(self) -> str:
        return _slot_label(self.slot_start)

    @property
    def end_label(self) -> str:
        return _slot_label(self.slot_end)


def _safe_story_excerpt(body: str, limit: int = 220) -> str:
    compact = " ".join(part.strip() for part in body.splitlines() if part.strip())
    if len(compact) <= limit:
        return compact
    return compact[: limit - 1].rstrip() + "…"


def _sheet_title(ws, title: str, subtitle: str, span_to: str) -> None:
    ws.merge_cells(f"A1:{span_to}1")
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color=PALETTE["white"])
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=PALETTE["navy"])
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"A2:{span_to}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, color=PALETTE["text_muted"])
    ws["A2"].fill = PatternFill(fill_type="solid", fgColor=PALETTE["slate_soft"])
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20


def _section_label(ws, cell_ref: str, text: str) -> None:
    cell = ws[cell_ref]
    cell.value = text
    cell.font = Font(size=11, bold=True, color=PALETTE["navy"])
    cell.fill = PatternFill(fill_type="solid", fgColor=PALETTE["teal_soft"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = THIN_BORDER


def _apply_card(ws, *, top_row: int, left_col: int, title: str, value: str, accent: str) -> None:
    value_cell = ws.cell(row=top_row, column=left_col)
    value_cell.value = value
    value_cell.font = Font(size=18, bold=True, color=PALETTE["navy"])
    value_cell.fill = PatternFill(fill_type="solid", fgColor=accent)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.border = THIN_BORDER

    label_cell = ws.cell(row=top_row + 1, column=left_col)
    label_cell.value = title
    label_cell.font = Font(size=10, color=PALETTE["text_muted"])
    label_cell.fill = PatternFill(fill_type="solid", fgColor=PALETTE["white"])
    label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    label_cell.border = THIN_BORDER

    ws.row_dimensions[top_row].height = 28
    ws.row_dimensions[top_row + 1].height = 22
    ws.column_dimensions[get_column_letter(left_col)].width = 18


def _add_table(ws, start_row: int, start_col: int, headers: list[str], rows: list[list[object]], name: str) -> None:
    for offset, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + offset)
        cell.value = header
        cell.font = Font(bold=True, color=PALETTE["white"])
        cell.fill = PatternFill(fill_type="solid", fgColor=PALETTE["teal"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_index, row_values in enumerate(rows, start=start_row + 1):
        for col_offset, value in enumerate(row_values):
            cell = ws.cell(row=row_index, column=start_col + col_offset)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

    last_row = start_row + max(len(rows), 1)
    last_col = start_col + len(headers) - 1
    table = Table(
        displayName=name,
        ref=f"{get_column_letter(start_col)}{start_row}:{get_column_letter(last_col)}{last_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _slot_label(slot_index: int) -> str:
    day = max(1, math.ceil(slot_index / HALF_DAY_SLOTS))
    half = "AM" if slot_index % HALF_DAY_SLOTS == 1 else "PM"
    return f"Day {day} {half}"


def _score_item(item: DeliveryItem) -> int:
    score = 50
    senior_rd_days = max(item.senior_rd_days, 0.5)

    if senior_rd_days <= 1:
        score += 8
    elif senior_rd_days <= 2:
        score += 4
    elif senior_rd_days >= 5:
        score -= 6

    if item.requirement_refs:
        score += min(6, len(item.requirement_refs) * 2)

    if item.requirement_source == "explicit":
        score += 6
    elif item.requirement_source == "inferred":
        score += 3
    else:
        score -= 4

    text = f"{item.title} {item.body}".lower()
    if any(keyword in text for keyword in BOOST_KEYWORDS):
        score += 8

    return score


def _priority_tier(score: int) -> str:
    for tier, minimum in PRIORITY_TIERS:
        if score >= minimum:
            return tier
    return PRIORITY_TIERS[-1][0]


def _schedule_bucket(tier: str) -> str:
    if tier == "P0 - Immediate":
        return "Now"
    if tier == "P1 - Next":
        return "Next"
    return "Later"


def _build_planning_rows(items: list[DeliveryItem]) -> list[PlanningRow]:
    sorted_items = sorted(
        items,
        key=lambda item: (
            PRIORITY_RANK[_priority_tier(_score_item(item))],
            -_score_item(item),
            item.senior_rd_days,
            item.group.lower(),
            item.title.lower(),
        ),
    )

    planning_rows: list[PlanningRow] = []
    current_slot = 1
    for index, item in enumerate(sorted_items, start=1):
        score = _score_item(item)
        tier = _priority_tier(score)
        duration_slots = max(1, int(round(max(item.senior_rd_days, 0.5) * HALF_DAY_SLOTS)))
        slot_start = current_slot
        slot_end = current_slot + duration_slots - 1
        current_slot = slot_end + 1

        planning_rows.append(
            PlanningRow(
                task_id=f"T-{index:03d}",
                priority_tier=tier,
                priority_score=score,
                scheduling_bucket=_schedule_bucket(tier),
                group=item.group,
                title=item.title,
                senior_rd_days=item.senior_rd_days,
                tracker_estimate=item.estimate,
                requirement_refs=item.requirement_refs,
                requirement_source=item.requirement_source,
                labels_text=", ".join(item.labels),
                source_story=_safe_story_excerpt(item.body),
                slot_start=slot_start,
                slot_end=slot_end,
            )
        )
    return planning_rows


def _populate_overview_sheet(ws, project_name: str, rows: list[PlanningRow], exported_at: str) -> None:
    ws.sheet_view.showGridLines = False
    _sheet_title(
        ws,
        f"{project_name} Delivery Plan",
        "Executive-ready planning workbook with prioritisation, traceability, and schedule views.",
        "H",
    )

    total_days = sum(row.senior_rd_days for row in rows)
    unmapped_count = sum(1 for row in rows if row.requirement_source == "unmapped")
    priority_now = sum(1 for row in rows if row.priority_tier == "P0 - Immediate")

    _apply_card(ws, top_row=4, left_col=1, title="Total Tasks", value=str(len(rows)), accent=PALETTE["teal_soft"])
    _apply_card(ws, top_row=4, left_col=2, title="Senior RD Days", value=f"{total_days:g}d", accent=PALETTE["gold_soft"])
    _apply_card(ws, top_row=4, left_col=3, title="Immediate Tasks", value=str(priority_now), accent=PALETTE["mint_soft"])
    _apply_card(ws, top_row=4, left_col=4, title="Unmapped Tasks", value=str(unmapped_count), accent=PALETTE["rose_soft"])

    _section_label(ws, "A8", "Planning Summary")
    ws["A9"] = "Project Name"
    ws["B9"] = project_name
    ws["A10"] = "Exported At"
    ws["B10"] = exported_at
    ws["A11"] = "Primary Planning Metric"
    ws["B11"] = "Senior RD ideal engineering days"
    ws["A12"] = "Tracker Compatibility"
    ws["B12"] = "Closest tracker estimate retained for Jira / GitHub publish flows"
    for row in range(9, 13):
        for column in ("A", "B"):
            cell = ws[f"{column}{row}"]
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column == "A":
                cell.font = Font(bold=True, color=PALETTE["navy"])
                cell.fill = PatternFill(fill_type="solid", fgColor=PALETTE["slate_soft"])

    _section_label(ws, "D8", "Priority Distribution")
    priority_rows = [
        [tier, sum(1 for row in rows if row.priority_tier == tier)]
        for tier, _ in PRIORITY_TIERS
    ]
    _add_table(ws, 9, 4, ["Priority Tier", "Tasks"], priority_rows, "PriorityDistribution")

    _section_label(ws, "G8", "Traceability Coverage")
    traceability_rows = [
        [label, sum(1 for row in rows if row.requirement_source == source)]
        for source, label in TRACEABILITY_LABEL.items()
    ]
    _add_table(ws, 9, 7, ["Traceability", "Tasks"], traceability_rows, "TraceabilityCoverage")

    _section_label(ws, "A15", "Requirement Coverage")
    coverage: dict[str, int] = {}
    for row in rows:
        for ref in row.requirement_refs:
            coverage[ref] = coverage.get(ref, 0) + 1
    coverage_rows = [[ref, count] for ref, count in sorted(coverage.items())]
    if not coverage_rows:
        coverage_rows = [["unmapped", len(rows)]]
    _add_table(ws, 16, 1, ["Requirement ID", "Tasks"], coverage_rows, "RequirementCoverage")

    _section_label(ws, "D15", "Schedule Assumptions")
    notes = [
        "One senior RD is the baseline implementer.",
        "Effort is shown in ideal engineering days, not elapsed calendar duration.",
        "Schedule view uses relative Day 1 / Day 2 slots to keep exports deterministic.",
        "Rows marked as Unmapped need traceability review before tracker publication.",
    ]
    for index, note in enumerate(notes, start=16):
        cell = ws[f"D{index}"]
        cell.value = f"• {note}"
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = THIN_BORDER
        ws.merge_cells(start_row=index, start_column=4, end_row=index, end_column=8)
        ws[f"D{index}"].fill = PatternFill(fill_type="solid", fgColor=PALETTE["white"])

    for column, width in {
        "A": 18, "B": 26, "C": 18, "D": 22, "E": 14, "F": 4, "G": 18, "H": 16
    }.items():
        ws.column_dimensions[column].width = width


def _populate_tasks_sheet(ws, project_name: str, rows: list[PlanningRow]) -> None:
    ws.sheet_view.showGridLines = False
    _sheet_title(
        ws,
        "Prioritised Tasks",
        f"Sorted by deterministic planning score for {project_name}.",
        "L",
    )

    headers = [
        "Task ID",
        "Priority Tier",
        "Priority Score",
        "Scheduling Bucket",
        "Senior RD Days",
        "Tracker Estimate",
        "Epic / Group",
        "Task",
        "Requirement IDs",
        "Traceability",
        "Labels",
        "Source Story",
    ]
    task_rows = [
        [
            row.task_id,
            row.priority_tier,
            row.priority_score,
            row.scheduling_bucket,
            row.senior_rd_days,
            row.tracker_estimate,
            row.group,
            row.title,
            row.requirement_refs_text,
            row.requirement_source_text,
            row.labels_text,
            row.source_story,
        ]
        for row in rows
    ]
    _add_table(ws, 4, 1, headers, task_rows, "PrioritisedTasks")

    for excel_row, plan_row in enumerate(rows, start=5):
        tier_cell = ws.cell(row=excel_row, column=2)
        tier_cell.fill = PatternFill(fill_type="solid", fgColor=PRIORITY_FILL[plan_row.priority_tier])
        tier_cell.font = Font(bold=True, color=PALETTE["navy"])

        traceability_cell = ws.cell(row=excel_row, column=10)
        traceability_cell.fill = PatternFill(fill_type="solid", fgColor=TRACEABILITY_FILL[plan_row.requirement_source])
        traceability_cell.font = Font(bold=True, color=PALETTE["navy"])

    ws.freeze_panes = "A5"
    for column, width in enumerate((11, 16, 14, 16, 14, 14, 18, 36, 24, 14, 22, 44), start=1):
        ws.column_dimensions[get_column_letter(column)].width = width


def _populate_schedule_sheet(ws, rows: list[PlanningRow]) -> None:
    ws.sheet_view.showGridLines = False
    _sheet_title(
        ws,
        "Gantt Schedule",
        "Relative Day / half-day view for one senior RD. This is a planning aid, not a calendar commitment.",
        "N",
    )

    max_slot = max((row.slot_end for row in rows), default=2)
    total_days = max(1, math.ceil(max_slot / HALF_DAY_SLOTS))

    ws["A4"] = "Task ID"
    ws["B4"] = "Priority"
    ws["C4"] = "Epic / Group"
    ws["D4"] = "Requirement IDs"
    ws["E4"] = "Start"
    ws["F4"] = "Finish"
    for cell_ref in ("A4", "B4", "C4", "D4", "E4", "F4"):
        cell = ws[cell_ref]
        cell.font = Font(bold=True, color=PALETTE["white"])
        cell.fill = PatternFill(fill_type="solid", fgColor=PALETTE["teal"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    timeline_start_col = 7
    for day_index in range(total_days):
        start_col = timeline_start_col + (day_index * HALF_DAY_SLOTS)
        end_col = start_col + HALF_DAY_SLOTS - 1
        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
        header_cell = ws.cell(row=3, column=start_col)
        header_cell.value = f"Day {day_index + 1}"
        header_cell.font = Font(bold=True, color=PALETTE["navy"])
        header_cell.fill = PatternFill(fill_type="solid", fgColor=PALETTE["slate_soft"])
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.border = THIN_BORDER

        for slot_offset, label in enumerate(("AM", "PM")):
            slot_cell = ws.cell(row=4, column=start_col + slot_offset)
            slot_cell.value = label
            slot_cell.font = Font(bold=True, color=PALETTE["white"])
            slot_cell.fill = PatternFill(fill_type="solid", fgColor=PALETTE["navy"])
            slot_cell.alignment = Alignment(horizontal="center", vertical="center")
            slot_cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(start_col + slot_offset)].width = 6

    for excel_row, plan_row in enumerate(rows, start=5):
        row_values = [
            plan_row.task_id,
            plan_row.priority_tier,
            plan_row.group,
            plan_row.requirement_refs_text,
            plan_row.start_label,
            plan_row.end_label,
        ]
        for column_index, value in enumerate(row_values, start=1):
            cell = ws.cell(row=excel_row, column=column_index)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

        ws.cell(row=excel_row, column=2).fill = PatternFill(
            fill_type="solid",
            fgColor=PRIORITY_FILL[plan_row.priority_tier],
        )

        for slot_index in range(plan_row.slot_start, plan_row.slot_end + 1):
            timeline_col = timeline_start_col + slot_index - 1
            cell = ws.cell(row=excel_row, column=timeline_col)
            cell.fill = PatternFill(fill_type="solid", fgColor=PRIORITY_FILL[plan_row.priority_tier])
            cell.border = THIN_BORDER
            cell.value = "■" if slot_index == plan_row.slot_start else ""
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "G5"
    for column, width in {"A": 11, "B": 16, "C": 18, "D": 24, "E": 12, "F": 12}.items():
        ws.column_dimensions[column].width = width


def _populate_guidance_sheet(ws, rows: list[PlanningRow]) -> None:
    ws.sheet_view.showGridLines = False
    _sheet_title(
        ws,
        "Scoring & Guidance",
        "Use this sheet to understand how ordering, effort, and traceability were derived.",
        "H",
    )

    _section_label(ws, "A4", "Priority Scoring Logic")
    scoring_rows = [
        ["Base score", "50 for every task before adjustments"],
        ["Low effort bonus", "+8 when senior RD effort is 1 day or less; +4 when it is 2 days or less"],
        ["High effort penalty", "-6 when senior RD effort is 5 days or more"],
        ["Requirement coverage bonus", "+2 per requirement ref, capped at +6"],
        ["Traceability bonus", "Explicit refs +6, inferred refs +3, unmapped refs -4"],
        ["Strategic keyword boost", f"+8 when matching terms such as {', '.join(BOOST_KEYWORDS[:6])}"],
    ]
    _add_table(ws, 5, 1, ["Signal", "Effect"], scoring_rows, "ScoringLogic")

    _section_label(ws, "E4", "Planning Guidance")
    guidance_rows = [
        ["Senior RD Estimate", "Ideal engineering days for one senior RD, rounded to 0.5 day increments"],
        ["Requirement IDs", "Primary traceability link back to the original requirement set"],
        ["Traceability", "Explicit = present in story, Inferred = matched from PRD, Unmapped = review needed"],
        ["Schedule view", "Relative Day 1 / Day 2 slots keep exports deterministic across runs"],
        ["Tracker Estimate", "Retained only for Jira / GitHub compatibility during the transition away from Story Points"],
    ]
    _add_table(ws, 5, 5, ["Field", "Meaning"], guidance_rows, "PlanningGuidance")

    _section_label(ws, "A13", "SWQA Review Checklist")
    checklist_rows = [
        ["Workbook rendering", "Open in Excel and Google Sheets, confirm widths, fills, frozen panes, and wrapped text"],
        ["Traceability", "Review every Unmapped or Inferred row before using the workbook as delivery evidence"],
        ["Schedule integrity", "Check that total senior RD days match the Gantt bars and overview totals"],
        ["Regression safety", "Confirm tracker publish previews still carry compatibility estimates and traceability notes"],
    ]
    _add_table(ws, 14, 1, ["Check", "Expected Evidence"], checklist_rows, "SwqaChecklist")

    _section_label(ws, "E13", "Current Export Snapshot")
    snapshot_rows = [
        ["Total tasks", len(rows)],
        ["Total senior RD days", f"{sum(row.senior_rd_days for row in rows):g}"],
        ["Explicit traceability", sum(1 for row in rows if row.requirement_source == "explicit")],
        ["Inferred traceability", sum(1 for row in rows if row.requirement_source == "inferred")],
        ["Unmapped traceability", sum(1 for row in rows if row.requirement_source == "unmapped")],
    ]
    _add_table(ws, 14, 5, ["Metric", "Value"], snapshot_rows, "ExportSnapshot")

    for column, width in {"A": 18, "B": 40, "C": 4, "D": 4, "E": 18, "F": 40, "G": 4, "H": 4}.items():
        ws.column_dimensions[column].width = width


def generate_delivery_excel(project_name: str, items: list[DeliveryItem]) -> io.BytesIO:
    """Generate a polished workbook for delivery planning."""
    workbook = Workbook()
    exported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    planning_rows = _build_planning_rows(items)

    overview_ws = workbook.active
    overview_ws.title = "Overview"
    _populate_overview_sheet(overview_ws, project_name, planning_rows, exported_at)

    tasks_ws = workbook.create_sheet("Prioritised Tasks")
    _populate_tasks_sheet(tasks_ws, project_name, planning_rows)

    schedule_ws = workbook.create_sheet("Gantt Schedule")
    _populate_schedule_sheet(schedule_ws, planning_rows)

    guidance_ws = workbook.create_sheet("Scoring & Guidance")
    _populate_guidance_sheet(guidance_ws, planning_rows)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
